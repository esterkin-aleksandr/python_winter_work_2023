import openpyxl
from openpyxl import Workbook
wb = Workbook()
# wb.save('Task_10.0_ex.xlsx')
wb = openpyxl.load_workbook('Task_10.0_ex.xlsx')
# print(wb.sheetnames)
lst=[]
for sheet in wb.sheetnames:
    ws = wb[sheet]
    n = ws.max_row * ws.max_column
    lst.append((sheet, n))
print(lst)
print(sorted(lst, key = lambda x: (x[0], -x[1])))