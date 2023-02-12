import openpyxl
from openpyxl import Workbook
wb = Workbook()
# wb.save('Task_10.1_ex.xlsx')
wb = openpyxl.load_workbook('Task_10.1_ex.xlsx')
# wb.create_sheet("Итог")
# wb.save('Task_10.1_ex.xlsx')
ws1=wb['Sheet']
ws2=wb['Итог']
sotrsum = {}
for i in range(1,ws1.max_row):
    sotr = ws1.cell(row=i, column=1).value
    sotrsum[sotr] = sotrsum.get(sotr, 0) + ws1.cell(row=i, column=2).value
print(sotrsum)
for j, k in sotrsum.items():
    ws2.append([j, k])
summ = 0
for x in range(1, ws2.max_row + 1):
    summ += int(ws2.cell(row=x, column=2).value)
ws2.append(['Общая сумма', summ])
wb.save('Task_10.1_ex.xlsx')